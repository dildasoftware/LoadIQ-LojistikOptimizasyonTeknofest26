"""
Ozet sayfasini gercek satir toplamina gore guncelle (yuvarlama farkini sifirla)
"""
import pandas as pd
import os
from openpyxl import load_workbook

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
pl_path = os.path.join(base_dir, "outputs", "Arac_Planlama.xlsx")

pl = pd.read_excel(pl_path, sheet_name="Arac Planlama")

kiralik = pl[pl['Arac Tipi'].str.startswith('Kiralik')]['Maliyet'].sum()
spot = pl[pl['Arac Tipi'].str.startswith('Spot')]['Maliyet'].sum()
toplam = pl['Maliyet'].sum()

print(f"Satir bazli hesaplanan degerler:")
print(f"  Kiralik : {kiralik:,.2f} TL")
print(f"  Spot    : {spot:,.2f} TL")
print(f"  Toplam  : {toplam:,.2f} TL")

# Ozet sayfasini guncelle
wb = load_workbook(pl_path)
ws = wb["Ozet"]

# Degerleri satir bazli toplamla guncelle
for row in ws.iter_rows():
    for cell in row:
        if cell.value == 'Toplam Kiralik Maliyet (TL)':
            ws.cell(row=cell.row, column=2).value = f"{kiralik:,.2f}"
        elif cell.value == 'Toplam Spot Maliyet (TL)':
            ws.cell(row=cell.row, column=2).value = f"{spot:,.2f}"
        elif cell.value == 'GENEL TOPLAM MALIYET (TL)':
            ws.cell(row=cell.row, column=2).value = f"{toplam:,.2f}"

wb.save(pl_path)
print(f"\nOzet sayfasi guncellendi: {pl_path}")
print("Artik satir toplami = Ozet toplami (0 fark)")
